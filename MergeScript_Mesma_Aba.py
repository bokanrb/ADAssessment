import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
import random

# Função para gerar cores hexadecimais aleatórias
def gerar_cor_aleatoria():
    cores = ['FF0000', '00FF00', '0000FF', 'FFFF00', 'FF00FF', '00FFFF', 'FFC0CB', '8A2BE2', 'A52A2A', '5F9EA0']
    return random.choice(cores)

# Defina o diretório onde seus arquivos CSV estão localizados
diretorio = 'C:\\Users\\leonardo.keller\\Downloads\\ADSecOutput\\ADSecOutput\\'

# Crie uma lista para armazenar os DataFrames
dfs = []

# Liste todos os arquivos CSV no diretório
for arquivo in os.listdir(diretorio):
    if arquivo.endswith('.csv'):
        caminho_arquivo = os.path.join(diretorio, arquivo)
        
        try:
            # Carregue o arquivo CSV em um DataFrame
            df = pd.read_csv(caminho_arquivo)
            
            # Verifique se o DataFrame está vazio (sem colunas)
            if not df.empty:
                # Adicione uma coluna em branco
                df.insert(0, '', '')  # Inserir coluna vazia no início
                dfs.append((df, arquivo))  # Armazenar DataFrame junto com o nome do arquivo
            else:
                print(f"Aviso: O arquivo '{arquivo}' está vazio e foi ignorado.")
        
        except pd.errors.EmptyDataError:
            print(f"Aviso: O arquivo '{arquivo}' não contém dados válidos e foi ignorado.")
        except Exception as e:
            print(f"Erro ao processar o arquivo '{arquivo}': {e}")

# Se houver DataFrames para combinar
if dfs:
    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active

    # Para cada DataFrame e nome de arquivo
    col_offset = 0  # Para controlar onde começar a escrever cada DataFrame
    for df, arquivo in dfs:
        # Função para remover a extensão do arquivo
        nome_arquivo_sem_extensao = os.path.splitext(arquivo)[0]

        # Adiciona o nome do arquivo como título (célula mesclada) sobre o cabeçalho
        ws.merge_cells(start_row=1, start_column=col_offset + 1, end_row=1, end_column=col_offset + len(df.columns))
        cell = ws.cell(row=1, column=col_offset + 1)
        cell.value = nome_arquivo_sem_extensao  # Coloca o nome do arquivo na célula mesclada
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinhamento centralizado
        
        # Aplica uma cor aleatória no cabeçalho
        cor_aleatoria = gerar_cor_aleatoria()
        fill = PatternFill(start_color=cor_aleatoria, end_color=cor_aleatoria, fill_type="solid")
        cell.fill = fill

        # Escreve os cabeçalhos e dados do DataFrame
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
            for c, value in enumerate(row, col_offset + 1):
                ws.cell(row=r, column=c, value=value)

        # Atualiza o deslocamento de colunas para o próximo DataFrame
        col_offset += len(df.columns) + 1  # Adiciona 1 para a coluna vazia entre os DataFrames

    # Salve o arquivo Excel
    wb.save('ADVsec.xlsx')
    print("Arquivo Excel gerado com sucesso: 'ADVsec.xlsx'")
else:
    print("Nenhum arquivo válido foi processado.")
