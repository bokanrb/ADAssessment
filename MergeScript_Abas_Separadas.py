import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Defina o diretório onde seus arquivos CSV estão localizados
diretorio = r'C:\\Users\\leonardo.keller\\Downloads\\ADSecOutput\\ADSecOutput'

# Lista para armazenar (DataFrame, nome_arquivo)
dfs = []

# Ler todos os arquivos CSV no diretório
for arquivo in os.listdir(diretorio):
    if arquivo.endswith('.csv'):
        caminho_arquivo = os.path.join(diretorio, arquivo)
        
        try:
            df = pd.read_csv(caminho_arquivo)
            
            # Verificar se o DataFrame está vazio
            if not df.empty:
                # (Opcional) Inserir uma coluna em branco na primeira posição
                df.insert(0, '', '')
                
                # Guardar o DataFrame junto com o nome do arquivo
                dfs.append((df, arquivo))
            else:
                print(f"Aviso: O arquivo '{arquivo}' está vazio.")
        
        except pd.errors.EmptyDataError:
            print(f"Aviso: O arquivo '{arquivo}' não contém dados válidos")
        except Exception as e:
            print(f"Erro ao processar o arquivo '{arquivo}': {e}")

# Se tivermos pelo menos um DataFrame, vamos criar o Excel
if dfs:
    # Cria um novo Workbook
    wb = Workbook()
    
    # Remover a primeira aba padrão criada automaticamente (opcional e preferível)
    wb.remove(wb.active)

    # Para cada (df, arquivo) na lista
    for df, arquivo in dfs:
        # Extrair o nome do arquivo sem extensão e informar quais arquivos estão sendo processados
        nome_arquivo_sem_extensao = os.path.splitext(arquivo)[0]
        print(f"Arquivo processado: {nome_arquivo_sem_extensao}")

        # Criar uma nova aba com o nome do arquivo (máximo 31 caracteres no Excel)
        ws = wb.create_sheet(title=nome_arquivo_sem_extensao[:31])
        
        # Escrever o DataFrame a partir da linha 1
        for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

    # Salva o Workbook em disco
    wb.save('resultado_abas_separadas.xlsx')
    print("Arquivo Excel gerado com sucesso: 'resultado_abas_separadas.xlsx'")
else:
    print("Nenhum arquivo válido foi processado.")
